from pathlib import Path

from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook

from inventario.models import (
    Categoria,
    DetalleEntrada,
    Entrada,
    Marca,
    Modelo,
    Producto,
    Usuario,
    Vehiculo,
)


class Command(BaseCommand):
    help = "Carga categorias y piezas del Excel en la base local de pruebas."

    def add_arguments(self, parser):
        parser.add_argument(
            "archivo",
            nargs="?",
            default="VENTAS.xlsx",
            help="Ruta del archivo .xlsx (por defecto: VENTAS.xlsx).",
        )
        parser.add_argument(
            "--limpiar",
            action="store_true",
            help="Elimina productos de prueba anteriores antes de cargar.",
        )

    @transaction.atomic
    def handle(self, *args, **options):
        archivo = Path(options["archivo"])
        if not archivo.exists():
            self.stderr.write(self.style.ERROR(f"No existe el archivo: {archivo}"))
            return

        if options["limpiar"]:
            Entrada.objects.filter(vehiculo__patente__startswith="PRUEBA").delete()
            Producto.objects.filter(nombre__startswith="[PRUEBA XLSX]").delete()
            Producto.objects.filter(nombre__in=self._nombres_excel(archivo)).delete()
            Vehiculo.objects.filter(patente__startswith="PRUEBA").delete()
            Modelo.objects.filter(nombre_modelo__in=["Yaris", "Fiesta", "Sail"]).delete()
            Marca.objects.filter(nombre_marca__in=["Toyota", "Ford", "Chevrolet"]).delete()

        workbook = load_workbook(archivo, read_only=True, data_only=True)
        hoja = workbook[workbook.sheetnames[0]]
        filas = [
            fila
            for fila in hoja.iter_rows(values_only=True)
            if any(valor is not None and str(valor).strip() for valor in fila)
        ]
        if not filas:
            self.stderr.write(self.style.ERROR("El Excel no contiene filas con datos."))
            return
        encabezados = [str(valor or "").strip().upper() for valor in filas[0]]
        usuario = Usuario.objects.filter(is_superuser=True).first()
        if not usuario:
            usuario = Usuario.objects.filter(is_staff=True).first()
        if not usuario:
            self.stderr.write(
                self.style.ERROR("Crea un superusuario antes de cargar datos de prueba.")
            )
            return

        entrada, _ = Entrada.objects.get_or_create(
            fecha="2026-09-05", vehiculo=None, usuario=usuario
        )
        creados = 0
        for indice, encabezado in enumerate(encabezados):
            if not encabezado or encabezado.startswith("COLUMNA"):
                continue
            categoria, _ = Categoria.objects.get_or_create(nombre_categoria=encabezado)
            for fila in filas[1:]:
                nombre = fila[indice] if indice < len(fila) else None
                if not nombre or not str(nombre).strip():
                    continue
                nombre = str(nombre).strip()
                producto, creado = Producto.objects.get_or_create(
                    nombre=nombre,
                    categoria=categoria,
                    vehiculo=None,
                    defaults={"costo": None},
                )
                DetalleEntrada.objects.get_or_create(
                    entrada=entrada, producto=producto, defaults={"cantidad": 3 + (creados % 4)}
                )
                creados += int(creado)

        self.stdout.write(self.style.SUCCESS(f"Carga completada: {creados} productos de prueba."))

    def _nombres_excel(self, archivo):
        workbook = load_workbook(archivo, read_only=True, data_only=True)
        hoja = workbook[workbook.sheetnames[0]]
        filas = [
            fila
            for fila in hoja.iter_rows(values_only=True)
            if any(valor is not None and str(valor).strip() for valor in fila)
        ]
        nombres = []
        for fila in filas[1:]:
            nombres.extend(str(valor).strip() for valor in fila if valor and str(valor).strip())
        return nombres