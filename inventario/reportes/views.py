from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Font

from ..pdf import rotacion_pdf_bytes, utilidad_pdf_bytes, ventas_pdf_bytes
from .queries import reporte_rotacion, reporte_utilidad, reporte_ventas, rango_desde_hasta


# ---------------------------------------------------------------------------
# Ventas por periodo
# ---------------------------------------------------------------------------
@login_required
def ventas(request):
    desde, hasta = rango_desde_hasta(request)
    datos = reporte_ventas(desde, hasta)
    contexto = {**datos, "desde": desde, "hasta": hasta}
    return render(request, "inventario/reportes/ventas.html", contexto)


@login_required
def ventas_exportar_pdf(request):
    desde, hasta = rango_desde_hasta(request)
    datos = reporte_ventas(desde, hasta)
    response = HttpResponse(
        ventas_pdf_bytes(datos, desde, hasta), content_type="application/pdf"
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_ventas.pdf"'
    return response


@login_required
def ventas_exportar_excel(request):
    desde, hasta = rango_desde_hasta(request)
    datos = reporte_ventas(desde, hasta)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"
    ws.append(["Fecha", "Documento", "Forma de pago", "Usuario", "Total"])
    for celda in ws[1]:
        celda.font = Font(bold=True)
    for venta in datos["ventas"]:
        ws.append(
            [
                venta.fecha_venta,
                str(venta.tipo_documento),
                str(venta.forma_pago),
                str(venta.usuario),
                venta.total_venta,
            ]
        )
    fila_total = ws.max_row + 1
    ws.cell(row=fila_total, column=4, value="Total").font = Font(bold=True)
    ws.cell(row=fila_total, column=5, value=datos["total_general"]).font = Font(bold=True)
    for fila in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for celda in fila:
            celda.number_format = "#,##0.00"
    for columna, ancho in {"A": 12, "B": 16, "C": 16, "D": 16, "E": 14}.items():
        ws.column_dimensions[columna].width = ancho

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_ventas.xlsx"'
    wb.save(response)
    return response


# ---------------------------------------------------------------------------
# Utilidad por mes
# ---------------------------------------------------------------------------
@login_required
def utilidad(request):
    desde, hasta = rango_desde_hasta(request, meses_por_defecto=6)
    datos = reporte_utilidad(desde, hasta)
    contexto = {**datos, "desde": desde, "hasta": hasta}
    return render(request, "inventario/reportes/utilidad.html", contexto)


@login_required
def utilidad_exportar_pdf(request):
    desde, hasta = rango_desde_hasta(request, meses_por_defecto=6)
    datos = reporte_utilidad(desde, hasta)
    response = HttpResponse(
        utilidad_pdf_bytes(datos, desde, hasta), content_type="application/pdf"
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_utilidad.pdf"'
    return response


@login_required
def utilidad_exportar_excel(request):
    desde, hasta = rango_desde_hasta(request, meses_por_defecto=6)
    datos = reporte_utilidad(desde, hasta)

    wb = Workbook()
    ws = wb.active
    ws.title = "Utilidad"
    ws.append(["Mes", "Ventas", "Gastos", "Utilidad"])
    for celda in ws[1]:
        celda.font = Font(bold=True)
    for fila in datos["filas"]:
        ws.append(
            [
                fila["mes"].strftime("%m-%Y"),
                fila["ventas"],
                fila["gastos"],
                fila["utilidad"],
            ]
        )
    fila_total = ws.max_row + 1
    ws.cell(row=fila_total, column=1, value="Total").font = Font(bold=True)
    ws.cell(row=fila_total, column=2, value=datos["total_ventas"]).font = Font(bold=True)
    ws.cell(row=fila_total, column=3, value=datos["total_gastos"]).font = Font(bold=True)
    ws.cell(row=fila_total, column=4, value=datos["total_utilidad"]).font = Font(bold=True)
    for fila in ws.iter_rows(min_row=2, min_col=2, max_col=4):
        for celda in fila:
            celda.number_format = "#,##0.00"
    for columna, ancho in {"A": 14, "B": 16, "C": 16, "D": 16}.items():
        ws.column_dimensions[columna].width = ancho

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_utilidad.xlsx"'
    wb.save(response)
    return response


# ---------------------------------------------------------------------------
# Rotacion de productos
# ---------------------------------------------------------------------------
@login_required
def rotacion(request):
    desde, hasta = rango_desde_hasta(request)
    datos = reporte_rotacion(desde, hasta)
    contexto = {**datos, "desde": desde, "hasta": hasta}
    return render(request, "inventario/reportes/rotacion.html", contexto)


@login_required
def rotacion_exportar_pdf(request):
    desde, hasta = rango_desde_hasta(request)
    datos = reporte_rotacion(desde, hasta)
    response = HttpResponse(
        rotacion_pdf_bytes(datos, desde, hasta), content_type="application/pdf"
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_rotacion.pdf"'
    return response


@login_required
def rotacion_exportar_excel(request):
    desde, hasta = rango_desde_hasta(request)
    datos = reporte_rotacion(desde, hasta)

    wb = Workbook()
    ws = wb.active
    ws.title = "Más vendidos"
    ws.append(["Producto", "Categoría", "Cantidad vendida"])
    for celda in ws[1]:
        celda.font = Font(bold=True)
    for producto in datos["mas_vendidos"]:
        ws.append([producto.nombre, str(producto.categoria), producto.cantidad_vendida])

    ws2 = wb.create_sheet("Menos vendidos")
    ws2.append(["Producto", "Categoría", "Cantidad vendida"])
    for celda in ws2[1]:
        celda.font = Font(bold=True)
    for producto in datos["menos_vendidos"]:
        ws2.append([producto.nombre, str(producto.categoria), producto.cantidad_vendida])

    for hoja in (ws, ws2):
        for columna, ancho in {"A": 28, "B": 18, "C": 18}.items():
            hoja.column_dimensions[columna].width = ancho

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_rotacion.xlsx"'
    wb.save(response)
    return response
