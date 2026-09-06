from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.db.models import F, Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from ..models import Categoria, Producto, Vehiculo
from ..servicios.importacion import COLUMNAS_MINIMAS, procesar_filas_productos
from ..servicios.inventario import productos_con_stock, valor_inventario
from .forms import (
    EdicionMasivaForm,
    ImportarProductosForm,
    InventarioFiltroForm,
    ProductoFiltroForm,
    ProductoForm,
)

ORDENES_VALIDOS = ("nombre", "-nombre", "categoria", "-categoria", "costo", "-costo")


def _productos_filtrados(request):
    form = InventarioFiltroForm(request.GET or None)
    productos = productos_con_stock()

    if form.is_valid():
        datos = form.cleaned_data
        if datos["categoria"]:
            productos = productos.filter(categoria=datos["categoria"])
        if datos["marca"]:
            productos = productos.filter(vehiculo__modelo__marca=datos["marca"])
        if datos["modelo"]:
            productos = productos.filter(vehiculo__modelo=datos["modelo"])
        if datos["estado"] == "disponible":
            productos = productos.filter(stock_disponible__gt=0)
        elif datos["estado"] == "agotado":
            productos = productos.filter(stock_disponible__lte=0)
    return form, productos


def _filtrar_lista_productos(request):
    filtro = ProductoFiltroForm(request.GET or None)
    productos = Producto.objects.filter(fecha_eliminacion__isnull=True).select_related(
        "categoria", "vehiculo__modelo__marca"
    )
    if filtro.is_valid():
        datos = filtro.cleaned_data
        if datos["nombre"]:
            productos = productos.filter(nombre__icontains=datos["nombre"])
        if datos["categoria"]:
            productos = productos.filter(
                categoria__nombre_categoria__icontains=datos["categoria"]
            )
        if datos["vehiculo"]:
            busqueda_vehiculo = datos["vehiculo"]
            productos = productos.filter(
                Q(vehiculo__patente__icontains=busqueda_vehiculo)
                | Q(vehiculo__modelo__nombre_modelo__icontains=busqueda_vehiculo)
                | Q(vehiculo__modelo__marca__nombre_marca__icontains=busqueda_vehiculo)
            )

    orden = request.GET.get("orden", "nombre")
    if orden not in ORDENES_VALIDOS:
        orden = "nombre"
    orden_sql = orden
    if orden in ("categoria", "-categoria"):
        orden_sql = f"{orden}__nombre_categoria"
    return filtro, productos.order_by(orden_sql, "nombre"), orden


@login_required
def inventario_visualizacion(request):
    form, productos = _productos_filtrados(request)

    resumen = productos.aggregate(
        productos=Sum("stock_disponible"),
        unidades_vendidas=Sum("total_vendido"),
        unidades_ingresadas=Sum("total_entradas"),
    )
    productos = list(productos)
    disponibles = sum(p.stock_disponible for p in productos)
    contexto = {
        "form": form,
        "productos": productos,
        "resumen": resumen,
        "valor_inventario": valor_inventario(productos),
        "metricas": [
            {"titulo": "Unidades disponibles", "valor": disponibles, "detalle": "Stock actual"},
            {"titulo": "Productos con stock", "valor": sum(p.stock_disponible > 0 for p in productos), "detalle": "Referencias activas"},
            {"titulo": "Unidades vendidas", "valor": sum(p.total_vendido for p in productos), "detalle": "Salidas registradas"},
        ],
    }
    return render(request, "inventario/visualizaciones/inventario.html", contexto)


@login_required
def inventario_valorizado(request):
    form, productos = _productos_filtrados(request)
    productos = list(productos)
    for producto in productos:
        producto.valor_stock = (producto.costo or 0) * producto.stock_disponible
    contexto = {
        "form": form,
        "productos": productos,
        "valor_inventario": valor_inventario(productos),
        "unidades_disponibles": sum(p.stock_disponible for p in productos),
        "productos_con_stock": sum(1 for p in productos if p.stock_disponible > 0),
        "metricas": [
            {"titulo": "Unidades disponibles", "valor": sum(p.stock_disponible for p in productos), "detalle": "Stock actual"},
            {"titulo": "Valor del inventario", "valor": f"${valor_inventario(productos):,.0f}", "detalle": "A costo de adquisición"},
            {"titulo": "Unidades vendidas", "valor": sum(p.total_vendido for p in productos), "detalle": "Salidas registradas"},
            {"titulo": "Productos con stock", "valor": sum(p.stock_disponible > 0 for p in productos), "detalle": "Referencias activas"},
        ],
    }
    return render(
        request, "inventario/visualizaciones/inventario_valorizado.html", contexto
    )


@login_required
def productos_lista(request):
    filtro, productos, orden_actual = _filtrar_lista_productos(request)
    return render(
        request,
        "inventario/visualizaciones/productos_lista.html",
        {
            "productos": productos,
            "filtro": filtro,
            "orden_actual": orden_actual,
            "form": ProductoForm(),
        },
    )


@login_required
def producto_crear(request):
    if request.method != "POST":
        return redirect("productos_lista")
    form = ProductoForm(request.POST)
    if form.is_valid():
        form.save()
        messages.success(request, "Producto creado correctamente.")
        return redirect("productos_lista")
    productos = Producto.objects.filter(fecha_eliminacion__isnull=True).select_related(
        "categoria", "vehiculo__modelo__marca"
    ).order_by("nombre")
    return render(
        request,
        "inventario/visualizaciones/productos_lista.html",
        {
            "productos": productos,
            "filtro": ProductoFiltroForm(request.GET or None),
            "form": form,
            "abrir_modal": True,
        },
    )


@login_required
def producto_editar(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    form = ProductoForm(request.POST or None, instance=producto)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Producto actualizado correctamente.")
        return redirect("productos_lista")
    plantilla = "inventario/visualizaciones/producto_form.html"
    if request.GET.get("partial"):
        plantilla = "inventario/visualizaciones/producto_form_modal.html"
    return render(request, plantilla, {"form": form, "producto": producto})


@login_required
def producto_eliminar(request, pk):
    if request.method != "POST":
        return redirect("productos_lista")
    if request.POST.get("confirmar_eliminacion") != "1":
        messages.error(request, "Debes confirmar la eliminación del producto.")
        return redirect("productos_lista")
    producto = get_object_or_404(
        Producto, pk=pk, fecha_eliminacion__isnull=True
    )
    producto.eliminar(request.user)
    messages.success(request, "Producto eliminado correctamente.")
    return redirect("productos_lista")


# ---------------------------------------------------------------------------
# Carga masiva por Excel
# ---------------------------------------------------------------------------
CABECERAS_EXPORT = [
    "codigo",
    "nombre",
    "categoria",
    "vehiculo_id",
    "vehiculo_desc",
    "costo",
    "precio_venta",
    "stock_actual",
    "id_producto",
]


@login_required
def productos_exportar_excel(request):
    _filtro, productos, _orden = _filtrar_lista_productos(request)
    stock_por_producto = {
        p.pk: p.stock_disponible for p in productos_con_stock()
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    ws.append(CABECERAS_EXPORT)
    for celda in ws[1]:
        celda.font = Font(bold=True)

    for producto in productos:
        ws.append(
            [
                producto.codigo or "",
                producto.nombre,
                producto.categoria.nombre_categoria,
                producto.vehiculo_id or "",
                str(producto.vehiculo) if producto.vehiculo_id else "",
                producto.costo,
                producto.precio_venta,
                stock_por_producto.get(producto.pk, 0),
                producto.pk,
            ]
        )

    ws_cat = wb.create_sheet("Categorías")
    ws_cat.append(["id", "nombre"])
    for celda in ws_cat[1]:
        celda.font = Font(bold=True)
    for categoria in Categoria.objects.order_by("nombre_categoria"):
        ws_cat.append([categoria.pk, categoria.nombre_categoria])

    ws_veh = wb.create_sheet("Vehículos")
    ws_veh.append(["id", "patente", "marca", "modelo", "anio"])
    for celda in ws_veh[1]:
        celda.font = Font(bold=True)
    for vehiculo in Vehiculo.objects.select_related("modelo__marca").order_by(
        "modelo__marca__nombre_marca", "modelo__nombre_modelo", "anio"
    ):
        ws_veh.append(
            [
                vehiculo.pk,
                vehiculo.patente or "",
                vehiculo.modelo.marca.nombre_marca,
                vehiculo.modelo.nombre_modelo,
                vehiculo.anio,
            ]
        )

    ws_ayuda = wb.create_sheet("Instrucciones")
    for fila in [
        ["CÓMO USAR ESTA PLANTILLA"],
        [],
        ["Columna", "¿Requerida?", "Detalle"],
        ["codigo", "No", "Vacío = producto nuevo (se genera PRD-000123). Con valor = se busca ese producto para actualizarlo."],
        ["nombre", "Sí (nuevos)", "Nombre de la pieza."],
        ["categoria", "Sí (nuevos)", "Nombre exacto. Si no existe, se crea. Ver hoja 'Categorías'."],
        ["vehiculo_id", "No", "ID de la hoja 'Vehículos'. Vacío = sin vehículo."],
        ["vehiculo_desc", "—", "Solo referencia, la importación lo ignora."],
        ["costo", "No", "Costo de adquisición. Vacío al actualizar = no se cambia."],
        ["precio_venta", "No", "Precio de venta. Vacío al actualizar = no se cambia."],
        ["stock_actual", "No", "Si es mayor al stock actual se crea una Entrada de ajuste por la diferencia. Menor = advertencia (hazlo con una venta)."],
        ["id_producto", "—", "Solo referencia, la importación lo ignora."],
    ]:
        ws_ayuda.append(fila)
    ws_ayuda["A1"].font = Font(bold=True, size=13)

    for hoja, anchos in {
        ws: [16, 30, 20, 12, 26, 12, 14, 12, 12],
        ws_cat: [8, 30],
        ws_veh: [8, 14, 16, 16, 8],
        ws_ayuda: [16, 14, 80],
    }.items():
        for indice, ancho in enumerate(anchos, start=1):
            hoja.column_dimensions[hoja.cell(row=1, column=indice).column_letter].width = ancho

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="productos.xlsx"'
    wb.save(response)
    return response


@login_required
def productos_importar(request):
    if request.method == "POST":
        form = ImportarProductosForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = form.cleaned_data["archivo"]
            archivo.seek(0)
            try:
                wb = load_workbook(archivo, read_only=True, data_only=True)
                hoja = wb["Productos"] if "Productos" in wb.sheetnames else wb.active
                filas_raw = [
                    fila
                    for fila in hoja.iter_rows(values_only=True)
                    if any(celda is not None and str(celda).strip() for celda in fila)
                ]
            except Exception as exc:  # noqa: BLE001
                messages.error(request, f"No se pudo leer el Excel: {exc}")
                return redirect("productos_importar")

            if not filas_raw:
                messages.error(request, "El archivo no tiene filas con datos.")
                return redirect("productos_importar")

            encabezados = [
                str(celda).strip() if celda is not None else "" for celda in filas_raw[0]
            ]
            faltantes = COLUMNAS_MINIMAS - set(encabezados)
            if faltantes:
                messages.error(
                    request, f"Faltan columnas obligatorias: {', '.join(sorted(faltantes))}."
                )
                return redirect("productos_importar")

            filas = [dict(zip(encabezados, fila)) for fila in filas_raw[1:]]
            creados, actualizados, ajustes, errores = procesar_filas_productos(
                filas, request.user
            )

            if creados or actualizados:
                messages.success(
                    request,
                    f"{creados} creado(s), {actualizados} actualizado(s), "
                    f"{ajustes} ajuste(s) de stock.",
                )
            for error in errores[:10]:
                messages.warning(request, error)
            if len(errores) > 10:
                messages.warning(request, f"... y {len(errores) - 10} aviso(s) más.")
            if not (creados or actualizados or errores):
                messages.info(request, "No se registraron cambios.")
            return redirect("productos_lista")
    else:
        form = ImportarProductosForm()
    return render(
        request, "inventario/visualizaciones/productos_importar.html", {"form": form}
    )


@login_required
def productos_edicion_masiva(request):
    accion = request.POST.get("accion") if request.method == "POST" else None
    form = EdicionMasivaForm(
        request.POST or None, exigir_cambios=(accion == "aplicar")
    )
    base = Producto.objects.filter(fecha_eliminacion__isnull=True).select_related(
        "categoria", "vehiculo__modelo__marca"
    )

    filtrados = base
    if form.is_valid():
        filtrados = form.filtrar(base)
    productos = filtrados.order_by("categoria__nombre_categoria", "nombre")

    if accion == "aplicar" and form.is_valid():
        datos = form.cleaned_data
        seleccion = request.POST.getlist("seleccion")
        objetivo = filtrados
        if not datos["aplicar_a_todos"]:
            if not seleccion:
                messages.error(request, "Selecciona al menos un producto (o marca «aplicar a todos»).")
                return render(
                    request,
                    "inventario/visualizaciones/productos_edicion_masiva.html",
                    {"form": form, "productos": productos},
                )
            objetivo = objetivo.filter(pk__in=seleccion)

        cambios = {}
        if datos["nueva_categoria"]:
            cambios["categoria"] = datos["nueva_categoria"]
        if datos["nuevo_vehiculo"]:
            cambios["vehiculo"] = datos["nuevo_vehiculo"]
        if datos["quitar_vehiculo"]:
            cambios["vehiculo"] = None
        if datos["nuevo_costo"] is not None:
            cambios["costo"] = datos["nuevo_costo"]
        if datos["nuevo_precio_venta"] is not None:
            cambios["precio_venta"] = datos["nuevo_precio_venta"]
        if datos["ajuste_costo_pct"] is not None:
            factor = Decimal("1") + datos["ajuste_costo_pct"] / Decimal("100")
            cambios["costo"] = F("costo") * factor
        if datos["ajuste_precio_pct"] is not None:
            factor = Decimal("1") + datos["ajuste_precio_pct"] / Decimal("100")
            cambios["precio_venta"] = F("precio_venta") * factor

        with transaction.atomic():
            afectados = objetivo.update(**cambios)
        messages.success(request, f"{afectados} producto(s) actualizado(s).")
        return redirect("productos_lista")

    return render(
        request,
        "inventario/visualizaciones/productos_edicion_masiva.html",
        {"form": form, "productos": productos},
    )
