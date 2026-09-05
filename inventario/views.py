from django.contrib import messages
from django.db.models import ProtectedError, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render
from django.urls import reverse_lazy
from django.views.generic import CreateView, DeleteView, ListView, UpdateView
from openpyxl import Workbook
from openpyxl.styles import Font

from .forms import ConceptoGastoForm, GastoForm
from .models import ConceptoGasto, Gasto
from .pdf import gasto_comprobante_pdf_bytes, gastos_pdf_bytes

from .forms import FormaPagoForm, RolForm, TipoDocumentoForm, UsuarioForm
from .models import FormaPago, Rol, TipoDocumento, Usuario
from .permisos import permiso_requerido, tiene_permiso


@login_required
def en_construccion(request, titulo):
    return render(request, "inventario/en_construccion.html", {"titulo": titulo})


def gastos_filtrados(request):
    """Gastos ordenados por fecha y filtrados por desde/hasta= si vienen
    en la URL. La usan la lista y las 2 vistas de exportacion, para que las
    tres apliquen el mismo filtro sin repetir la logica."""
    queryset = Gasto.objects.select_related(
        "concepto", "forma_pago", "usuario", "tipo_documento"
    ).order_by("-fecha")
    desde = request.GET.get("desde")
    hasta = request.GET.get("hasta")
    if desde:
        queryset = queryset.filter(fecha__gte=desde)
    if hasta:
        queryset = queryset.filter(fecha__lte=hasta)
    return queryset


class GastoListView(ListView):
    model = Gasto
    context_object_name = "gastos"

    def get_queryset(self):
        return gastos_filtrados(self.request)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["total"] = self.get_queryset().aggregate(total=Sum("monto"))["total"]
        context["desde"] = self.request.GET.get("desde", "")
        context["hasta"] = self.request.GET.get("hasta", "")
        return context


def gastos_exportar_pdf(request):
    gastos = gastos_filtrados(request)
    total = gastos.aggregate(total=Sum("monto"))["total"] or 0

    response = HttpResponse(gastos_pdf_bytes(gastos, total), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="gastos.pdf"'
    return response



def gastos_exportar_excel(request):
    gastos = gastos_filtrados(request)
    total = gastos.aggregate(total=Sum("monto"))["total"] or 0

    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos"

    encabezados = [
        "Fecha", "Concepto", "Forma de pago", "Monto",
        "Tipo Doc.", "N.° Documento", "Observaciones", "Usuario",
    ]
    ws.append(encabezados)
    for celda in ws[1]:
        celda.font = Font(bold=True)

    for gasto in gastos:
        ws.append(
            [
                gasto.fecha,
                str(gasto.concepto),
                str(gasto.forma_pago),
                gasto.monto,
                str(gasto.tipo_documento) if gasto.tipo_documento else "",
                gasto.numero_documento or "",
                gasto.observaciones or "",
                str(gasto.usuario),
            ]
        )

    fila_total = ws.max_row + 1
    ws.cell(row=fila_total, column=3, value="Total").font = Font(bold=True)
    celda_total = ws.cell(row=fila_total, column=4, value=total)
    celda_total.font = Font(bold=True)

    # Formato numerico de Excel (no texto): separa miles/decimales segun el
    # locale del Excel que lo abra, y se puede seguir sumando/ordenado.
    for fila in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for celda in fila:
            celda.number_format = "#,##0.00"

    anchos = {"A": 12, "B": 20, "C": 16, "D": 14, "E": 12, "F": 16, "G": 30, "H": 16}
    for columna, ancho in anchos.items():
        ws.column_dimensions[columna].width = ancho

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="gastos.xlsx"'
    wb.save(response)
    return response


class GastoFormMixin:
    model = Gasto
    form_class = GastoForm
    template_name = "inventario/gasto_form.html"
    success_url = reverse_lazy("gastos")


class GastoCreateView(GastoFormMixin, CreateView):
    def form_valid(self, form):
        form.instance.usuario = self.request.user
        messages.success(self.request, "Gasto registrado correctamente.")
        return super().form_valid(form)


class GastoUpdateView(GastoFormMixin, UpdateView):
    def form_valid(self, form):
        messages.success(self.request, "Gasto actualizado correctamente.")
        return super().form_valid(form)


class GastoDeleteView(DeleteView):
    model = Gasto
    success_url = reverse_lazy("gastos")

    def form_valid(self, form):
        messages.success(self.request, "Gasto eliminado correctamente.")
        return super().form_valid(form)


def gasto_comprobante(request, pk):
    gasto = get_object_or_404(
        Gasto.objects.select_related("concepto", "forma_pago", "usuario", "tipo_documento"), pk=pk
    )
    return render(request, "inventario/gasto_comprobante.html", {"gasto": gasto})


def gasto_comprobante_pdf(request, pk):
    gasto = get_object_or_404(
        Gasto.objects.select_related("concepto", "forma_pago", "usuario", "tipo_documento"), pk=pk
    )
    response = HttpResponse(
        gasto_comprobante_pdf_bytes(gasto), content_type="application/pdf"
    )
    response["Content-Disposition"] = f'attachment; filename="comprobante_gasto_{gasto.pk}.pdf"'
    return response


class ConceptoGastoListView(ListView):
    model = ConceptoGasto
    ordering = "nombre_gasto"
    context_object_name = "conceptos"
    template_name = "inventario/concepto_list.html"


class ConceptoGastoFormMixin:
    model = ConceptoGasto
    form_class = ConceptoGastoForm
    template_name = "inventario/concepto_form.html"
    success_url = reverse_lazy("conceptos")


class ConceptoGastoCreateView(ConceptoGastoFormMixin, CreateView):
    def form_valid(self, form):
        messages.success(self.request, "Concepto registrado correctamente.")
        return super().form_valid(form)


class ConceptoGastoUpdateView(ConceptoGastoFormMixin, UpdateView):
    def form_valid(self, form):
        messages.success(self.request, "Concepto actualizado correctamente.")
        return super().form_valid(form)


class ConceptoGastoDeleteView(DeleteView):
    model = ConceptoGasto
    template_name = "inventario/concepto_confirm_delete.html"
    success_url = reverse_lazy("conceptos")

    def form_valid(self, form):
        try:
            response = super().form_valid(form)
        except ProtectedError:
            messages.error(
                self.request,
                "No se puede eliminar: hay gastos registrados con ese concepto.",
            )
            return self.get(self.request, *self.args, **self.kwargs)
        messages.success(self.request, "Concepto eliminado correctamente.")
        return response
